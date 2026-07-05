# Reads all sheets from an existing Excel file, prints cell values, and duplicates each sheet

from openpyxl import load_workbook


def read_spreadsheet(sheet):
    max_rows = sheet.max_row
    max_cols = sheet.max_column

    for row in range(1, max_rows + 1):
        for col in range(1, max_cols + 1):
            print(sheet.cell(row=row, column=col).value, end=' | ')
        print('\n')


workbook = load_workbook(filename='Example.xlsx', read_only=False)
sheet_names = workbook.sheetnames

for name in sheet_names:
    sheet = workbook[name]
    print('Sheet -', sheet.title, ':\n')
    read_spreadsheet(sheet)

for name in sheet_names:
    sheet = workbook[name]
    workbook.copy_worksheet(sheet)

workbook.save('Example.xlsx')
