# Reads an Excel file to calculate total population, city count, and average population per city
# Note: replace 'filename.xlsx' and 'column_name' with your actual file and column names

from openpyxl import load_workbook


def population_count():
    wb = load_workbook(filename='filename.xlsx', read_only=True)
    sheet = wb.active
    max_row = sheet.max_row
    max_col = sheet.max_column
    total_population = 0

    for col in range(1, max_col + 1):
        if sheet.cell(row=1, column=col).value == 'column_name':
            for row in range(2, max_row + 1):
                total_population += sheet.cell(row=row, column=col).value

    print('Total population: %3.f' % total_population)
    return total_population


def city_count():
    wb = load_workbook(filename='filename.xlsx', read_only=True)
    sheet = wb.active
    max_row = sheet.max_row
    max_col = sheet.max_column
    total_cities = 0

    for col in range(1, max_col + 1):
        if sheet.cell(row=1, column=col).value == 'column_name':
            for row in range(2, max_row + 1):  # fix: was reusing variable 'b' from outer loop
                if sheet.cell(row=row, column=col).value is not None:  # fix: was checking the method, not the value
                    total_cities += 1

    print('Total cities:', total_cities)
    return total_cities


def population_average(total_population, total_cities):
    average = int(total_population) / total_cities
    print('Average population per city: {}'.format(int(average)))


population_average(population_count(), city_count())
