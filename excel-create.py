# Creates a new Excel workbook with multiple sheets, adds data and a formula, and saves the file

from openpyxl import Workbook

workbook = Workbook()

sheet1 = workbook.active
sheet1.title = 'Sheet01'

sheet3 = workbook.create_sheet('Sheet03')
sheet2 = workbook.create_sheet('Sheet02', 1)
sheet4 = workbook.create_sheet('Sheet04')

sheet4['A1'] = 'Name 1'
sheet4['A2'] = 'Name 2'
sheet4['A3'] = 'Name 3'
sheet4['B1'] = 18
sheet4['B2'] = 14
sheet4['B3'] = 16
sheet4['B4'] = '=AVERAGE(B1:B3)'

values = [(1,), (2,), (3,)]

for row in values:  # fix: was all on one line
    sheet2.append(row)

sheet1.cell(row=2, column=2, value='HELLO')

sheet5 = workbook.create_sheet('Sheet05')

print(workbook.sheetnames)
workbook.save('ExampleFile01.xlsx')
